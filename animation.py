import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import openpyxl

# Load the sequence of moves from two Excel files
wb1 = openpyxl.load_workbook('data/test.xlsx')
wb2 = openpyxl.load_workbook('data/joint_angles_neg_good.xlsx')

sheet1 = wb1.active
sheet2 = wb2.active

sequence_of_moves1 = []
sequence_of_moves2 = []

# Read every second row starting from the second row, and columns 3 to 6 for the first file
for i, row in enumerate(sheet1.iter_rows(min_row=2, min_col=3, max_col=6, values_only=True)):
    cleaned_row = []
    for value in row:
        if isinstance(value, str):
            value = value.replace('−', '-').replace(',', '.')
        cleaned_row.append(float(value))
    sequence_of_moves1.append(cleaned_row)

# Read every second row starting from the second row, and columns 3 to 6 for the second file
for i, row in enumerate(sheet2.iter_rows(min_row=2, min_col=3, max_col=6, values_only=True)):
    cleaned_row = []
    for value in row:
        if isinstance(value, str):
            value = value.replace('−', '-').replace(',', '.')
        cleaned_row.append(float(value))
    sequence_of_moves2.append(cleaned_row)

# Function to compute the forward kinematics
def forward_kinematics(q1, q2, q3, q4):
    L1, L2, L3, L4 = 0.50, 0.93, 0.93, 0.50  # Link lengths
    x0, y0, z0 = 0, 0, 0
    x1, y1, z1 = 0, 0, L1
    x2, y2, z2 = L2 * np.cos(q1) * np.cos(q2), L2 * np.sin(q1) * np.cos(q2), L1 + L2 * np.sin(q2)
    x3, y3, z3 = x2 + L3 * np.cos(q1) * np.cos(q2 + q3), y2 + L3 * np.sin(q1) * np.cos(q2 + q3), z2 + L3 * np.sin(q2 + q3)
    x4, y4, z4 = x3 + L4 * np.cos(q1) * np.cos(q2 + q3 + q4), y3 + L4 * np.sin(q1) * np.cos(q2 + q3 + q4), z3 + L4 * np.sin(q2 + q3 + q4)
    return np.array([[x0, y0, z0], [x1, y1, z1], [x2, y2, z2], [x3, y3, z3], [x4, y4, z4]])

# Create the animation
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')

# Set equal scaling
ax.set_box_aspect([1, 1, 1])  # Aspect ratio is 1:1:1

# Set limits
ax.set_xlim([-2, 2])
ax.set_ylim([-2, 2])
ax.set_zlim([-2, 2])

line1, = ax.plot([], [], [], 'o-', lw=2, color='blue', label='Dataset 1')
line2, = ax.plot([], [], [], 'o-', lw=2, color='red', label='Dataset 2')
scatter1 = ax.scatter([], [], [], color='blue', s=10)
scatter2 = ax.scatter([], [], [], color='red', s=10)

end_effector_positions1 = []
end_effector_positions2 = []

def init():
    line1.set_data([], [])
    line1.set_3d_properties([])
    line2.set_data([], [])
    line2.set_3d_properties([])
    scatter1._offsets3d = ([], [], [])
    scatter2._offsets3d = ([], [], [])
    return line1, line2, scatter1, scatter2

def animate(i):
    angles1 = sequence_of_moves1[i]
    angles2 = sequence_of_moves2[i]
    positions1 = forward_kinematics(*angles1)
    positions2 = forward_kinematics(*angles2)
    line1.set_data(positions1[:, 0], positions1[:, 1])
    line1.set_3d_properties(positions1[:, 2])
    line2.set_data(positions2[:, 0], positions2[:, 1])
    line2.set_3d_properties(positions2[:, 2])
    
    end_effector_positions1.append(positions1[-1])
    end_effector_positions2.append(positions2[-1])
    
    scatter1._offsets3d = (np.array(end_effector_positions1)[:, 0], np.array(end_effector_positions1)[:, 1], np.array(end_effector_positions1)[:, 2])
    scatter2._offsets3d = (np.array(end_effector_positions2)[:, 0], np.array(end_effector_positions2)[:, 1], np.array(end_effector_positions2)[:, 2])
    
    return line1, line2, scatter1, scatter2

ani = animation.FuncAnimation(fig, animate, init_func=init, frames=min(len(sequence_of_moves1), len(sequence_of_moves2)), interval=200, blit=True)

plt.legend()
plt.show()